import xlrd
from xlrd.sheet import ctype_text
import openpyxl
import pickle
#>>> wb = openpyxl.load_workbook('example.xlsx') 
workbook = openpyxl.load_workbook('convertcsv.xlsx')

worksheet = workbook.get_sheet_by_name('Sheet 1')
ncols = worksheet.max_column
nrows = worksheet.max_row

columns_dict = {}

for c in range(1,ncols+1):
	columns_dict[c]=worksheet.cell(row=1, column=c).value
#print columns_dict


 


#num_cols = worksheet.ncols   # Number of columns

final_dict = {}
#for row_idx in range(0, worksheet.nrows):    # Iterate 
#print type(worksheet.cell(row=1, column=2).value)
#print worksheet.max_row
for r in range(2,nrows+1):
	temp_dict  = {}
	if worksheet.cell(row=r,column=2).value :
		temp_dict[worksheet.cell(row=r,column=2).value] = worksheet.cell(row=r,column=3).value

	if worksheet.cell(row=r,column=11).value :
		temp_dict[worksheet.cell(row=r,column=11).value] = worksheet.cell(row=r,column=12).value

	

	if worksheet.cell(row=r,column=15).value :
		temp_dict[worksheet.cell(row=r,column=15).value] = worksheet.cell(row=r,column=16).value
	
		
	if worksheet.cell(row=r,column=19).value :
		temp_dict[worksheet.cell(row=r,column=19).value] = worksheet.cell(row=r,column=20).value
	

	if worksheet.cell(row=r,column=23).value :
		temp_dict[worksheet.cell(row=r,column=23).value] = worksheet.cell(row=r,column=24).value
	

	if worksheet.cell(row=r,column=27).value :
		#print worksheet.cell(row=r,column=27).value
		temp_dict[worksheet.cell(row=r,column=27).value] = worksheet.cell(row=r,column=28).value
	
		
	if worksheet.cell(row=r,column=31).value :
		temp_dict[worksheet.cell(row=r,column=31).value] = worksheet.cell(row=r,column=32).value	
	

	if worksheet.cell(row=r,column=37).value :
		temp_dict[worksheet.cell(row=r,column=37).value] = worksheet.cell(row=r,column=38).value
	
		
	if worksheet.cell(row=r,column=43).value :
		temp_dict[worksheet.cell(row=r,column=43).value] = worksheet.cell(row=r,column=44).value
	
	#print temp_dict

	final_dict[worksheet.cell(row=r,column=8).value] = temp_dict

pickle.dump(final_dict,open('actual_id_aspects_pol_dump.p','wb'))


print final_dict